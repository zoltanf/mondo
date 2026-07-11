"""`mondo export` command group — dump a board to CSV / TSV / JSON / XLSX / MD / HTML / PDF.

Uses cursor pagination (reused from `item list`) plus a one-shot column-title
fetch for readable headers. monday's API has no native board-export endpoint —
every format is rendered client-side. Formats:

- csv / tsv: RFC 4180, one row per item, columns = meta + one-per-board-column.
  Cells starting with = + - @ (or tab/CR) get a `'` prefix so spreadsheets
  don't execute them as formulas; `--no-sanitize-formulas` disables the guard
  and `mondo import board` strips it back off.
- json: list of objects keyed by column id + metadata
- xlsx: one styled sheet for items (+ one for subitems if --include-subitems)
- md / html / pdf: human-readable; grouped per monday group by default
  (one section per group), or a single flat table with `--flat`. html reuses
  the doc stylesheet; pdf renders that HTML through WeasyPrint.

Output defaults to stdout for csv/tsv/json/md/html, and requires `--out` for
xlsx and pdf (binary / file formats).
"""

from __future__ import annotations

import csv
import html
import io
import json
from collections import Counter
from enum import StrEnum
from pathlib import Path
from typing import TYPE_CHECKING, Any

import typer

from mondo.api.errors import MondoError, NotFoundError, UsageError
from mondo.api.pagination import MAX_PAGE_SIZE, iter_items_page
from mondo.api.queries import (
    ITEMS_PAGE_INITIAL_WITH_SUBITEMS,
    ITEMS_PAGE_NEXT_WITH_SUBITEMS,
)
from mondo.cli._examples import epilog_for
from mondo.cli._exec import client_or_exit, handle_mondo_error_or_exit, usage_error_or_exit
from mondo.cli._pdf import render_pdf
from mondo.cli.context import GlobalOpts
from mondo.docs import _HTML_STYLE
from mondo.domain.column_cache import fetch_board_columns
from mondo.domain.resolve import resolve_required_id
from mondo.services.items import build_query_params, split_filter_expr
from mondo.util.sanitize import guard_formula

if TYPE_CHECKING:
    from mondo.api.client import MondayClient

app = typer.Typer(
    no_args_is_help=True,
    context_settings={"help_option_names": ["-h", "--help"]},
)


class ExportFormat(StrEnum):
    csv = "csv"
    tsv = "tsv"
    json = "json"
    xlsx = "xlsx"
    md = "md"
    html = "html"
    pdf = "pdf"


META_FIELDS = ("id", "name", "state", "group")
# Formats that render a board-name title and (by default) per-group sections.
_GROUPED_FORMATS = (ExportFormat.md, ExportFormat.html, ExportFormat.pdf)
# Binary / file-only formats that cannot stream to stdout (require --out).
_BINARY_FORMATS = (ExportFormat.xlsx, ExportFormat.pdf)

BOARD_NAME_QUERY = "query ($board: ID!) { boards(ids: [$board]) { name } }"


# ----- helpers -----


def _fetch_columns(opts: GlobalOpts, client: MondayClient, board_id: int) -> list[dict[str, Any]]:
    """Return the board's full (non-archived) column defs, in display order.

    Full defs (with `settings_str`) are kept so `build_query_params` can resolve
    `--filter` labels; the renderers only read `id`/`title`.
    """
    try:
        cols = fetch_board_columns(client, board_id, store=opts.columns_cache_store(board_id))
    except NotFoundError:
        raise typer.Exit(code=6) from None
    return [c for c in cols if not c.get("archived")]


def _fetch_board_name(client: MondayClient, board_id: int) -> str:
    """Best-effort board name for the md/html/pdf title.

    Never a hard error: any failure or empty result falls back to `Board <id>`.
    """
    try:
        result = client.execute(BOARD_NAME_QUERY, {"board": board_id})
    except MondoError:
        return f"Board {board_id}"
    boards = (result.get("data") or {}).get("boards") or []
    first = boards[0] if boards else None
    name = first.get("name") if isinstance(first, dict) else None
    return name or f"Board {board_id}"


def _resolve_columns(columns: list[dict[str, Any]], spec: str) -> list[dict[str, Any]]:
    """Project `columns` to the `--columns` subset, preserving requested order.

    Each comma-separated token matches by column id (exact) or by title
    (case-insensitive). A title shared by several columns selects them all. An
    unmatched token raises `UsageError` naming it.
    """
    by_id = {str(c.get("id")).lower(): c for c in columns}
    by_title: dict[str, list[dict[str, Any]]] = {}
    for c in columns:
        by_title.setdefault(str(c.get("title")).lower(), []).append(c)
    out: list[dict[str, Any]] = []
    for raw in spec.split(","):
        tok = raw.strip()
        if not tok:
            continue
        key = tok.lower()
        matches = [by_id[key]] if key in by_id else by_title.get(key, [])
        if not matches:
            raise UsageError(f"--columns: no column matches {tok!r}.")
        for col in matches:
            if col not in out:
                out.append(col)
    if not out:
        raise UsageError("--columns expects a comma-separated list of column ids.")
    return out


def _column_labels(columns: list[dict[str, Any]]) -> list[str]:
    """Display label per column, made unique so rows can be keyed by it.

    Rows are dicts keyed by these labels, so a label must not collide with
    another column's or with a meta field (`id`/`name`/`state`/`group`) — else
    one column's data would silently overwrite another's. monday allows
    duplicate column titles, so any title that repeats, or that shadows a meta
    field, is suffixed with its (unique) column id. Unique, non-shadowing titles
    are kept verbatim, so the common case is unchanged.
    """
    titles = [str(c.get("title")) for c in columns]
    counts = Counter(titles)
    labels: list[str] = []
    for col, title in zip(columns, titles, strict=True):
        if counts[title] > 1 or title in META_FIELDS:
            labels.append(f"{title} ({col.get('id')})")
        else:
            labels.append(title)
    return labels


def _group_key(item: dict[str, Any]) -> tuple[str | None, str]:
    """(group id, group title) for an item — id buckets sections, title displays."""
    group = item.get("group") or {}
    return (group.get("id"), group.get("title") or "")


def _column_text(item: dict[str, Any], col: dict[str, Any]) -> str:
    for cv in item.get("column_values") or []:
        if cv.get("id") == col["id"]:
            # Text-first so the CSV round-trip stays stable (rating/checkbox
            # export their raw `text`, not glyphs). mirror/board_relation
            # return a null `text`, so fall back to their `display_value`.
            return cv.get("text") or cv.get("display_value") or ""
    return ""


def _item_row(
    item: dict[str, Any], columns: list[dict[str, Any]], labels: list[str]
) -> dict[str, Any]:
    """Flat dict: meta fields + one key per column (keyed by its display label)."""
    row: dict[str, Any] = {
        "id": item.get("id"),
        "name": item.get("name"),
        "state": item.get("state"),
        "group": (item.get("group") or {}).get("title"),
    }
    for col, label in zip(columns, labels, strict=True):
        row[label] = _column_text(item, col)
    return row


def _subitem_row(
    subitem: dict[str, Any],
    parent: dict[str, Any],
    columns: list[dict[str, Any]],
    labels: list[str],
) -> dict[str, Any]:
    row: dict[str, Any] = {
        "parent_item_id": parent.get("id"),
        "parent_item_name": parent.get("name"),
        "id": subitem.get("id"),
        "name": subitem.get("name"),
        "state": subitem.get("state"),
    }
    for col, label in zip(columns, labels, strict=True):
        row[label] = _column_text(subitem, col)
    return row


# ----- format writers -----


def _write_csv(
    rows: list[dict[str, Any]],
    headers: list[str],
    delimiter: str,
    stream: Any,
    *,
    sanitize: bool,
) -> None:
    writer = csv.DictWriter(
        stream,
        fieldnames=headers,
        delimiter=delimiter,
        quoting=csv.QUOTE_MINIMAL,
        lineterminator="\n",
        extrasaction="ignore",
    )

    def cell(value: Any) -> Any:
        # Board data (item names, column text, column titles) is controlled
        # by anyone with board access — guard formula-looking cells so
        # opening the export in a spreadsheet can't execute them.
        return guard_formula(value) if sanitize else value

    writer.writerow({h: cell(h) for h in headers})
    for row in rows:
        writer.writerow({h: cell(row.get(h, "")) for h in headers})


def _write_json(payload: Any, stream: Any) -> None:
    json.dump(payload, stream, indent=2, ensure_ascii=False)
    stream.write("\n")


def _md_table(rows: list[dict[str, Any]], headers: list[str], stream: Any) -> None:
    def esc(v: Any) -> str:
        return str(v if v is not None else "").replace("|", "\\|").replace("\n", " ")

    stream.write("| " + " | ".join(esc(h) for h in headers) + " |\n")
    stream.write("|" + "|".join(["---"] * len(headers)) + "|\n")
    for row in rows:
        stream.write("| " + " | ".join(esc(row.get(h, "")) for h in headers) + " |\n")


def _write_markdown(
    item_rows: list[dict[str, Any]],
    col_labels: list[str],
    group_keys: list[tuple[str | None, str]],
    subitem_rows: list[dict[str, Any]] | None,
    subitem_headers: list[str] | None,
    board_name: str,
    flat: bool,
    stream: Any,
) -> None:
    stream.write(f"# {board_name}\n\n")
    if flat:
        _md_table(item_rows, [*META_FIELDS, *col_labels], stream)
    else:
        headers = ["id", "name", "state", *col_labels]
        for title, group_rows in _group_rows(item_rows, group_keys):
            stream.write(f"## {title}\n\n")
            _md_table(group_rows, headers, stream)
            stream.write("\n")
    if subitem_rows is not None and subitem_headers is not None:
        # Leading blank line: in flat mode the items table ends flush against
        # this heading, and GFM requires a blank line before a heading.
        stream.write("\n## Subitems\n\n")
        _md_table(subitem_rows, subitem_headers, stream)


def _html_table(rows: list[dict[str, Any]], headers: list[str], stream: Any) -> None:
    def esc(v: Any) -> str:
        return html.escape(str(v if v is not None else ""))

    stream.write("<table>\n<thead>\n<tr>")
    stream.write("".join(f"<th>{esc(h)}</th>" for h in headers))
    stream.write("</tr>\n</thead>\n<tbody>\n")
    for row in rows:
        stream.write("<tr>")
        stream.write("".join(f"<td>{esc(row.get(h, ''))}</td>" for h in headers))
        stream.write("</tr>\n")
    stream.write("</tbody>\n</table>\n")


def _render_board_html(
    item_rows: list[dict[str, Any]],
    col_labels: list[str],
    group_keys: list[tuple[str | None, str]],
    subitem_rows: list[dict[str, Any]] | None,
    subitem_headers: list[str] | None,
    board_name: str,
    flat: bool,
) -> str:
    body = io.StringIO()
    body.write(f'<h1 class="doc-title">{html.escape(board_name)}</h1>\n')
    if flat:
        _html_table(item_rows, [*META_FIELDS, *col_labels], body)
    else:
        headers = ["id", "name", "state", *col_labels]
        for title, group_rows in _group_rows(item_rows, group_keys):
            body.write(f"<h2>{html.escape(title)}</h2>\n")
            _html_table(group_rows, headers, body)
    if subitem_rows is not None and subitem_headers is not None:
        body.write("<h2>Subitems</h2>\n")
        _html_table(subitem_rows, subitem_headers, body)
    return (
        "<!DOCTYPE html>\n"
        '<html lang="en">\n<head>\n<meta charset="utf-8">\n'
        '<meta name="viewport" content="width=device-width, initial-scale=1">\n'
        f"<title>{html.escape(board_name)}</title>\n"
        f"<style>\n{_HTML_STYLE}</style>\n</head>\n<body>\n"
        f"{body.getvalue()}</body>\n</html>\n"
    )


def _group_rows(
    item_rows: list[dict[str, Any]], group_keys: list[tuple[str | None, str]]
) -> list[tuple[str, list[dict[str, Any]]]]:
    """Bucket rows by group id (in first-seen / board order), yielding (title, rows).

    Bucketing on the stable group id keeps two distinct groups that share a
    title as separate sections; the title is carried only for the heading. Rows
    with no group id fall back to bucketing by title.
    """
    order: list[Any] = []
    buckets: dict[Any, tuple[str, list[dict[str, Any]]]] = {}
    for row, (gid, gtitle) in zip(item_rows, group_keys, strict=True):
        key = gid if gid is not None else ("title", gtitle)
        if key not in buckets:
            buckets[key] = (gtitle, [])
            order.append(key)
        buckets[key][1].append(row)
    return [buckets[k] for k in order]


def _write_xlsx(
    items_rows: list[dict[str, Any]],
    items_headers: list[str],
    subitems_rows: list[dict[str, Any]] | None,
    subitems_headers: list[str] | None,
    out_path: Path,
) -> None:
    from openpyxl import Workbook  # type: ignore[import-untyped]
    from openpyxl.styles import Font  # type: ignore[import-untyped]
    from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

    def fill_sheet(ws: Any, headers: list[str], rows: list[dict[str, Any]]) -> None:
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for idx, header in enumerate(headers, start=1):
            longest = max(
                (len(str(row.get(header, ""))) for row in rows),
                default=0,
            )
            width = min(max(len(header), longest) + 2, 60)
            ws.column_dimensions[get_column_letter(idx)].width = width

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "items"
    fill_sheet(ws, items_headers, items_rows)
    if subitems_rows is not None and subitems_headers is not None:
        fill_sheet(wb.create_sheet("subitems"), subitems_headers, subitems_rows)
    wb.save(out_path)


# ----- command -----


@app.command("board", epilog=epilog_for("export board"))
def board_cmd(
    ctx: typer.Context,
    board_pos: int | None = typer.Argument(
        None, metavar="[BOARD_ID]", help="Board ID (positional)."
    ),
    board_flag: int | None = typer.Option(None, "--board", help="Board ID (flag form)."),
    fmt: ExportFormat = typer.Option(
        ExportFormat.csv,
        "--format",
        "-f",
        help="Output format: csv, tsv, json, xlsx, md, html, pdf.",
        case_sensitive=False,
    ),
    out: Path | None = typer.Option(
        None,
        "--out",
        help="Output file path. Required for xlsx and pdf. Omit to write to stdout "
        "(csv/tsv/json/md/html).",
    ),
    flat: bool = typer.Option(
        False,
        "--flat",
        help="md/html/pdf: emit a single flat table (with a `group` column) "
        "instead of one section per group. No effect on csv/tsv/json/xlsx.",
    ),
    group_id: str | None = typer.Option(
        None,
        "--group",
        help="Restrict to one group (alias for --filter group=<id>).",
        rich_help_panel="Filters",
    ),
    filter_expr: list[str] | None = typer.Option(
        None,
        "--filter",
        help="Server-side filter rule like 'status=Done' or 'status!=Stuck' (repeatable).",
        rich_help_panel="Filters",
    ),
    columns_sel: str | None = typer.Option(
        None,
        "--columns",
        metavar="COL1,COL2",
        help="Project to these columns only (by id or title, case-insensitive). "
        "Meta columns (id/name/state/group) are always kept.",
        rich_help_panel="Filters",
    ),
    include_subitems: bool = typer.Option(
        False, "--include-subitems", help="Also export each item's subitems."
    ),
    sanitize_formulas: bool = typer.Option(
        True,
        "--sanitize-formulas/--no-sanitize-formulas",
        help="csv/tsv: prefix cells starting with = + - @ (or tab/CR) with a "
        "single quote so spreadsheets don't execute them as formulas. "
        "`mondo import board` strips the prefix back off on re-import.",
    ),
    limit: int = typer.Option(MAX_PAGE_SIZE, "--limit", help=f"Page size (max {MAX_PAGE_SIZE})."),
    max_items: int | None = typer.Option(
        None, "--max-items", help="Stop after this many items total."
    ),
) -> None:
    """Export a board's items (and optionally subitems) in the chosen format."""
    opts: GlobalOpts = ctx.ensure_object(GlobalOpts)
    board_id = resolve_required_id(board_pos, board_flag, flag_name="--board", resource="board")

    if fmt in _BINARY_FORMATS and out is None:
        usage_error_or_exit(f"{fmt.value} is a binary format; pass --out PATH.")

    # --group is sugar over --filter group=<id>; merge it in.
    if group_id is not None:
        filter_expr = [*(filter_expr or []), f"group={group_id}"]

    # Fail fast on malformed --filter before opening the client (usage = exit 2).
    parsed_filters: list[tuple[str, str, str]] = []
    if filter_expr:
        try:
            parsed_filters = [split_filter_expr(f) for f in filter_expr]
        except UsageError as e:
            usage_error_or_exit(str(e))

    # Preflight WeasyPrint after local arg validation but before any network
    # work: a malformed --filter still surfaces first, and a first-time user
    # without WeasyPrint gets the install hint without paying for a board fetch
    # (matches the doc-export PDF path).
    if fmt is ExportFormat.pdf:
        from mondo.cli._pdf import find_weasyprint, install_hint

        if find_weasyprint() is None:
            handle_mondo_error_or_exit(MondoError(install_hint()))

    client = client_or_exit(opts)
    try:
        with client:
            columns = _fetch_columns(opts, client, board_id)

            # Build --filter params from the FULL column set: label resolution
            # must work even when --columns projects the filtered column away.
            column_defs = {c["id"]: c for c in columns}
            query_params = build_query_params(parsed_filters, None, column_defs, board_id=board_id)

            # --columns narrows only what's rendered, not what's filtered on.
            render_columns = columns
            if columns_sel is not None:
                try:
                    render_columns = _resolve_columns(columns, columns_sel)
                except UsageError as e:
                    usage_error_or_exit(str(e))

            board_name = _fetch_board_name(client, board_id) if fmt in _GROUPED_FORMATS else ""

            query_initial = None
            query_next = None
            if include_subitems:
                query_initial = ITEMS_PAGE_INITIAL_WITH_SUBITEMS
                query_next = ITEMS_PAGE_NEXT_WITH_SUBITEMS
            iterator_kwargs: dict[str, Any] = {
                "board_id": board_id,
                "limit": limit,
                "max_items": max_items,
            }
            if query_params is not None:
                iterator_kwargs["query_params"] = query_params
            if query_initial is not None and query_next is not None:
                iterator_kwargs["query_initial"] = query_initial
                iterator_kwargs["query_next"] = query_next

            items = list(iter_items_page(client, **iterator_kwargs))
    except MondoError as e:
        handle_mondo_error_or_exit(e)

    col_labels = _column_labels(render_columns)
    item_rows = [_item_row(it, render_columns, col_labels) for it in items]
    group_keys = [_group_key(it) for it in items]

    subitem_headers: list[str] | None = None
    subitem_rows: list[dict[str, Any]] | None = None
    if include_subitems:
        subitem_headers = ["parent_item_id", "parent_item_name", "id", "name", "state", *col_labels]
        subitem_rows = []
        for parent in items:
            for sub in parent.get("subitems") or []:
                subitem_rows.append(_subitem_row(sub, parent, render_columns, col_labels))

    _dispatch(
        fmt,
        item_rows,
        col_labels,
        group_keys,
        subitem_rows,
        subitem_headers,
        board_name,
        flat,
        out,
        sanitize_formulas,
    )


def _dispatch(
    fmt: ExportFormat,
    item_rows: list[dict[str, Any]],
    col_labels: list[str],
    group_keys: list[tuple[str | None, str]],
    subitem_rows: list[dict[str, Any]] | None,
    subitem_headers: list[str] | None,
    board_name: str,
    flat: bool,
    out: Path | None,
    sanitize_formulas: bool,
) -> None:
    item_headers = [*META_FIELDS, *col_labels]
    if fmt is ExportFormat.xlsx:
        assert out is not None
        _write_xlsx(item_rows, item_headers, subitem_rows, subitem_headers, out)
        return

    if fmt is ExportFormat.pdf:
        assert out is not None
        html_text = _render_board_html(
            item_rows, col_labels, group_keys, subitem_rows, subitem_headers, board_name, flat
        )
        try:
            render_pdf(html_text, out)
        except MondoError as e:
            handle_mondo_error_or_exit(e)
        return

    buf: Any = out.open("w", encoding="utf-8", newline="") if out is not None else io.StringIO()

    try:
        if fmt in (ExportFormat.csv, ExportFormat.tsv):
            delim = "," if fmt is ExportFormat.csv else "\t"
            _write_csv(item_rows, item_headers, delim, buf, sanitize=sanitize_formulas)
            if subitem_rows is not None and subitem_headers is not None:
                buf.write("\n")
                _write_csv(subitem_rows, subitem_headers, delim, buf, sanitize=sanitize_formulas)
        elif fmt is ExportFormat.json:
            payload: dict[str, Any] = {"items": item_rows}
            if subitem_rows is not None:
                payload["subitems"] = subitem_rows
            _write_json(payload, buf)
        elif fmt is ExportFormat.md:
            _write_markdown(
                item_rows,
                col_labels,
                group_keys,
                subitem_rows,
                subitem_headers,
                board_name,
                flat,
                buf,
            )
        elif fmt is ExportFormat.html:
            buf.write(
                _render_board_html(
                    item_rows,
                    col_labels,
                    group_keys,
                    subitem_rows,
                    subitem_headers,
                    board_name,
                    flat,
                )
            )
    finally:
        if out is not None:
            buf.close()
        else:
            typer.echo(buf.getvalue(), nl=False)
