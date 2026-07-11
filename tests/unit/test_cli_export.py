"""End-to-end CLI tests for `mondo export board ...`."""

from __future__ import annotations

import csv
import io
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook
from pytest_httpx import HTTPXMock
from typer.testing import CliRunner

from mondo.cli.main import app

ENDPOINT = "https://api.monday.com/v2"
runner = CliRunner()


@pytest.fixture(autouse=True)
def _clean_env(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    monkeypatch.delenv("MONDO_PROFILE", raising=False)
    monkeypatch.delenv("MONDAY_API_VERSION", raising=False)
    monkeypatch.setenv("MONDO_CONFIG", str(tmp_path / "nope.yaml"))
    monkeypatch.setenv("MONDAY_API_TOKEN", "env-token-abcdef-long-enough")
    # Isolate from the user's real cache dir and disable caching by default
    # — tests exercise the live fetch paths unless they explicitly opt in.
    monkeypatch.setenv("MONDO_CACHE_DIR", str(tmp_path / "cache"))
    monkeypatch.setenv("MONDO_CACHE_ENABLED", "false")


def _ok(data: dict) -> dict:
    return {"data": data, "extensions": {"request_id": "r"}}


COLS_RESPONSE = _ok(
    {
        "boards": [
            {
                "id": "42",
                "name": "B",
                "columns": [
                    {"id": "status", "title": "Status", "type": "status", "archived": False},
                    {"id": "date4", "title": "Due", "type": "date", "archived": False},
                    {
                        "id": "archived_col",
                        "title": "Old",
                        "type": "text",
                        "archived": True,
                    },
                ],
            }
        ]
    }
)


def _items_page(items: list[dict], cursor: str | None = None) -> dict:
    return _ok(
        {
            "boards": [
                {"items_page": {"cursor": cursor, "items": items}},
            ]
        }
    )


def _item(id_: str, name: str, cols: dict[str, str], group: str = "topics") -> dict:
    return {
        "id": id_,
        "name": name,
        "state": "active",
        "group": {"id": group, "title": group.title()},
        "column_values": [
            {"id": cid, "type": "text", "text": val, "value": None} for cid, val in cols.items()
        ],
    }


def _board_name(name: str = "My Board") -> dict:
    return _ok({"boards": [{"name": name}]})


# --- format tests ---


class TestCsvExport:
    def test_basic(self, httpx_mock: HTTPXMock) -> None:
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page(
                [
                    _item("1", "Alpha", {"status": "Done", "date4": "2026-04-25"}),
                    _item("2", "Beta", {"status": "Working on it", "date4": ""}),
                ]
            ),
        )
        result = runner.invoke(app, ["export", "board", "--board", "42", "--format", "csv"])
        assert result.exit_code == 0, result.stdout
        rows = list(csv.reader(io.StringIO(result.stdout)))
        assert rows[0] == ["id", "name", "state", "group", "Status", "Due"]
        assert rows[1][4] == "Done"
        assert rows[2][0] == "2"

    def test_omits_archived_columns(self, httpx_mock: HTTPXMock) -> None:
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        httpx_mock.add_response(
            url=ENDPOINT, method="POST", json=_items_page([_item("1", "X", {})])
        )
        result = runner.invoke(app, ["export", "board", "--board", "42", "-f", "csv"])
        assert result.exit_code == 0, result.stdout
        # "Old" is archived — should not appear in headers
        first_line = result.stdout.splitlines()[0]
        assert "Old" not in first_line

    def test_tsv_delimiter(self, httpx_mock: HTTPXMock) -> None:
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        httpx_mock.add_response(
            url=ENDPOINT, method="POST", json=_items_page([_item("1", "A", {})])
        )
        result = runner.invoke(app, ["export", "board", "--board", "42", "-f", "tsv"])
        assert result.exit_code == 0, result.stdout
        assert "\t" in result.stdout.splitlines()[0]

    def test_writes_to_file(self, httpx_mock: HTTPXMock, tmp_path: Path) -> None:
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page([_item("1", "A", {"status": "Done"})]),
        )
        out = tmp_path / "board.csv"
        result = runner.invoke(
            app,
            ["export", "board", "--board", "42", "-f", "csv", "--out", str(out)],
        )
        assert result.exit_code == 0, result.stdout
        content = out.read_text()
        assert "id,name,state,group,Status,Due" in content


class TestJsonExport:
    def test_shape(self, httpx_mock: HTTPXMock) -> None:
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page([_item("1", "A", {"status": "Done"})]),
        )
        result = runner.invoke(app, ["export", "board", "--board", "42", "-f", "json"])
        assert result.exit_code == 0, result.stdout
        parsed = json.loads(result.stdout)
        assert "items" in parsed
        assert parsed["items"][0]["Status"] == "Done"
        assert "subitems" not in parsed


class TestMarkdownExport:
    def test_flat_pipe_table(self, httpx_mock: HTTPXMock) -> None:
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=_board_name())
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page([_item("1", "A", {"status": "Done"})]),
        )
        result = runner.invoke(app, ["export", "board", "--board", "42", "-f", "md", "--flat"])
        assert result.exit_code == 0, result.stdout
        # Flat mode: a single table with a `group` column under the board title.
        # The title is kept; only the per-group `##` sections are dropped.
        assert result.stdout.startswith("# My Board\n")
        assert "| id | name | state | group | Status | Due |" in result.stdout
        assert "|---|" in result.stdout
        assert "## " not in result.stdout

    def test_grouped_is_default(self, httpx_mock: HTTPXMock) -> None:
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=_board_name("Roadmap"))
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page(
                [
                    _item("1", "A", {"status": "Done"}, group="topics"),
                    _item("2", "B", {"status": "Working"}, group="next"),
                ]
            ),
        )
        result = runner.invoke(app, ["export", "board", "--board", "42", "-f", "md"])
        assert result.exit_code == 0, result.stdout
        out = result.stdout
        # Board-name title + one section per group; `group` is the heading, not
        # a table column.
        assert out.startswith("# Roadmap\n")
        assert "## Topics" in out
        assert "## Next" in out
        assert "| id | name | state | Status | Due |" in out
        assert "group" not in out.splitlines()[2]

    def test_board_name_fallback(self, httpx_mock: HTTPXMock) -> None:
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        # Empty boards result -> fall back to "Board <id>", never a hard error.
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=_ok({"boards": []}))
        httpx_mock.add_response(
            url=ENDPOINT, method="POST", json=_items_page([_item("1", "A", {})])
        )
        result = runner.invoke(app, ["export", "board", "--board", "42", "-f", "md"])
        assert result.exit_code == 0, result.stdout
        assert result.stdout.startswith("# Board 42\n")

    def test_subitems_section_has_blank_line_separator(self, httpx_mock: HTTPXMock) -> None:
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=_board_name())
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page(
                [
                    {
                        **_item("1", "Parent", {"status": "Done"}),
                        "subitems": [
                            {
                                "id": "10",
                                "name": "Child A",
                                "state": "active",
                                "column_values": [
                                    {"id": "status", "text": "In progress", "type": "status"}
                                ],
                            }
                        ],
                    }
                ]
            ),
        )
        result = runner.invoke(
            app,
            ["export", "board", "--board", "42", "-f", "md", "--flat", "--include-subitems"],
        )
        assert result.exit_code == 0, result.stdout
        out = result.stdout
        # Flat items table ends flush against the heading; GFM needs the blank line.
        assert "\n\n## Subitems\n" in out
        assert "Child A" in out

    def test_table_header_escapes_pipe(self, httpx_mock: HTTPXMock) -> None:
        cols = _ok(
            {
                "boards": [
                    {
                        "id": "42",
                        "name": "B",
                        "columns": [
                            {"id": "c1", "title": "Cost | EUR", "type": "text", "archived": False}
                        ],
                    }
                ]
            }
        )
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=cols)
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=_board_name())
        httpx_mock.add_response(
            url=ENDPOINT, method="POST", json=_items_page([_item("1", "A", {"c1": "5"})])
        )
        result = runner.invoke(app, ["export", "board", "--board", "42", "-f", "md", "--flat"])
        assert result.exit_code == 0, result.stdout
        # A pipe in a column title must be escaped so the GFM table stays aligned.
        assert "Cost \\| EUR" in result.stdout


class TestHtmlExport:
    def test_title_groups_and_rows(self, httpx_mock: HTTPXMock) -> None:
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=_board_name("Roadmap"))
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page(
                [
                    _item("1", "Alpha", {"status": "Done"}, group="topics"),
                    _item("2", "Beta", {"status": "Working"}, group="next"),
                ]
            ),
        )
        result = runner.invoke(app, ["export", "board", "--board", "42", "-f", "html"])
        assert result.exit_code == 0, result.stdout
        out = result.stdout
        assert "<!DOCTYPE html>" in out
        # Reuses the doc stylesheet (it styles <table>, has @page A4).
        assert "@page" in out
        assert "<h1" in out and "Roadmap" in out
        assert "<h2>Topics</h2>" in out
        assert "<h2>Next</h2>" in out
        assert "<table>" in out
        assert "<td>Alpha</td>" in out

    def test_escaping(self, httpx_mock: HTTPXMock) -> None:
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=_board_name("A & B"))
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page([_item("1", "<script>", {"status": "x & y"})]),
        )
        result = runner.invoke(app, ["export", "board", "--board", "42", "-f", "html"])
        assert result.exit_code == 0, result.stdout
        out = result.stdout
        assert "<script>" not in out.replace("<script>", "")  # sanity
        assert "&lt;script&gt;" in out
        assert "x &amp; y" in out
        assert "A &amp; B" in out

    def test_flat_single_table(self, httpx_mock: HTTPXMock) -> None:
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=_board_name())
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page(
                [
                    _item("1", "A", {}, group="topics"),
                    _item("2", "B", {}, group="next"),
                ]
            ),
        )
        result = runner.invoke(app, ["export", "board", "--board", "42", "-f", "html", "--flat"])
        assert result.exit_code == 0, result.stdout
        out = result.stdout
        assert "<h2>" not in out
        assert out.count("<table>") == 1
        assert "<th>group</th>" in out


class TestPdfExport:
    def test_requires_out(self, httpx_mock: HTTPXMock) -> None:
        result = runner.invoke(app, ["export", "board", "--board", "42", "-f", "pdf"])
        assert result.exit_code == 2
        assert httpx_mock.get_requests() == []

    def test_calls_render_pdf_with_board_html(
        self, httpx_mock: HTTPXMock, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=_board_name("Roadmap"))
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page([_item("1", "Alpha", {"status": "Done"})]),
        )
        captured: dict[str, object] = {}

        def fake_render_pdf(html_text: str, out: Path) -> None:
            captured["html"] = html_text
            captured["out"] = out
            out.write_bytes(b"%PDF-1.4 fake")

        import mondo.cli.export as export_mod

        monkeypatch.setattr(export_mod, "render_pdf", fake_render_pdf)
        monkeypatch.setattr("mondo.cli._pdf.find_weasyprint", lambda: "/usr/bin/weasyprint")
        out = tmp_path / "board.pdf"
        result = runner.invoke(
            app,
            ["export", "board", "--board", "42", "-f", "pdf", "--out", str(out)],
        )
        assert result.exit_code == 0, result.stdout
        assert captured["out"] == out
        html_text = captured["html"]
        assert isinstance(html_text, str)
        assert "<!DOCTYPE html>" in html_text
        assert "Roadmap" in html_text
        assert "<td>Alpha</td>" in html_text

    def test_missing_weasyprint_fails_before_fetch(
        self, httpx_mock: HTTPXMock, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        # No WeasyPrint on PATH: fail fast with the install hint before any
        # board fetch, so the user isn't made to paginate first.
        monkeypatch.setattr("mondo.cli._pdf.find_weasyprint", lambda: None)
        out = tmp_path / "board.pdf"
        result = runner.invoke(
            app,
            ["export", "board", "--board", "42", "-f", "pdf", "--out", str(out)],
        )
        assert result.exit_code != 0
        assert httpx_mock.get_requests() == []
        assert "weasyprint" in (result.stdout + str(result.exception or "")).lower()

    def test_malformed_filter_beats_weasyprint_preflight(
        self, httpx_mock: HTTPXMock, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        # With WeasyPrint absent, a malformed --filter is a local usage error and
        # must surface before the install hint (both are pre-network).
        monkeypatch.setattr("mondo.cli._pdf.find_weasyprint", lambda: None)
        out = tmp_path / "board.pdf"
        result = runner.invoke(
            app,
            ["export", "board", "--board", "42", "-f", "pdf", "--out", str(out), "--filter", "bad"],
        )
        assert result.exit_code == 2
        assert httpx_mock.get_requests() == []
        assert "weasyprint" not in (result.stdout + str(result.exception or "")).lower()


class TestXlsxExport:
    def test_requires_out(self, httpx_mock: HTTPXMock) -> None:
        result = runner.invoke(app, ["export", "board", "--board", "42", "-f", "xlsx"])
        assert result.exit_code == 2
        assert httpx_mock.get_requests() == []

    def test_writes_workbook(self, httpx_mock: HTTPXMock, tmp_path: Path) -> None:
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page([_item("1", "A", {"status": "Done"})]),
        )
        out = tmp_path / "board.xlsx"
        result = runner.invoke(
            app,
            ["export", "board", "--board", "42", "-f", "xlsx", "--out", str(out)],
        )
        assert result.exit_code == 0, result.stdout
        wb = load_workbook(out)
        ws = wb["items"]
        headers = [c.value for c in ws[1]]
        assert headers == ["id", "name", "state", "group", "Status", "Due"]
        row2 = [c.value for c in ws[2]]
        assert row2[0] == "1"
        assert row2[4] == "Done"


# --- subitems ---


class TestIncludeSubitems:
    def test_uses_with_subitems_query(self, httpx_mock: HTTPXMock) -> None:
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page(
                [
                    {
                        **_item("1", "Parent", {"status": "Done"}),
                        "subitems": [
                            {
                                "id": "10",
                                "name": "Child A",
                                "state": "active",
                                "column_values": [
                                    {"id": "status", "text": "In progress", "type": "status"}
                                ],
                            }
                        ],
                    }
                ]
            ),
        )
        result = runner.invoke(
            app,
            [
                "export",
                "board",
                "--board",
                "42",
                "-f",
                "json",
                "--include-subitems",
            ],
        )
        assert result.exit_code == 0, result.stdout
        # Items-page request must be the subitems variant
        items_body = json.loads(httpx_mock.get_requests()[-1].content)
        assert "subitems" in items_body["query"]
        parsed = json.loads(result.stdout)
        assert parsed["subitems"][0]["parent_item_id"] == "1"
        assert parsed["subitems"][0]["Status"] == "In progress"

    def test_xlsx_emits_subitems_sheet(self, httpx_mock: HTTPXMock, tmp_path: Path) -> None:
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page(
                [
                    {
                        **_item("1", "P", {}),
                        "subitems": [
                            {
                                "id": "10",
                                "name": "Sub",
                                "state": "active",
                                "column_values": [],
                            }
                        ],
                    }
                ]
            ),
        )
        out = tmp_path / "b.xlsx"
        result = runner.invoke(
            app,
            [
                "export",
                "board",
                "--board",
                "42",
                "-f",
                "xlsx",
                "--out",
                str(out),
                "--include-subitems",
            ],
        )
        assert result.exit_code == 0, result.stdout
        wb = load_workbook(out)
        assert "items" in wb.sheetnames
        assert "subitems" in wb.sheetnames


# --- pagination ---


class TestPagination:
    def test_follows_cursor(self, httpx_mock: HTTPXMock) -> None:
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page([_item("1", "A", {})], cursor="C"),
        )
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_ok({"next_items_page": {"cursor": None, "items": [_item("2", "B", {})]}}),
        )
        result = runner.invoke(app, ["export", "board", "--board", "42", "-f", "json"])
        assert result.exit_code == 0, result.stdout
        parsed = json.loads(result.stdout)
        assert [i["id"] for i in parsed["items"]] == ["1", "2"]


# --- xlsx polish ---


class TestXlsxStyle:
    def test_freeze_filter_and_bold(self, httpx_mock: HTTPXMock, tmp_path: Path) -> None:
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page([_item("1", "A", {"status": "Done"})]),
        )
        out = tmp_path / "board.xlsx"
        result = runner.invoke(
            app, ["export", "board", "--board", "42", "-f", "xlsx", "--out", str(out)]
        )
        assert result.exit_code == 0, result.stdout
        wb = load_workbook(out)
        ws = wb["items"]
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref == ws.dimensions
        assert ws["A1"].font.bold is True


# --- filtering & column subset ---


_COLS_WITH_STATUS_LABELS = _ok(
    {
        "boards": [
            {
                "id": "42",
                "name": "B",
                "columns": [
                    {
                        "id": "status",
                        "title": "Status",
                        "type": "status",
                        "archived": False,
                        "settings_str": json.dumps({"labels": {"1": "Done"}}),
                    },
                    {"id": "date4", "title": "Due", "type": "date", "archived": False},
                ],
            }
        ]
    }
)


class TestFilter:
    def test_filter_carries_query_params(self, httpx_mock: HTTPXMock) -> None:
        # settings_str labels let the status codec resolve "Done" -> index, proving
        # the full column defs (not just {id,title,type}) reach build_query_params.
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=_COLS_WITH_STATUS_LABELS)
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page([_item("1", "A", {"status": "Done"})]),
        )
        result = runner.invoke(
            app,
            ["export", "board", "--board", "42", "-f", "csv", "--filter", "status=Done"],
        )
        assert result.exit_code == 0, result.stdout
        items_body = json.loads(httpx_mock.get_requests()[-1].content)
        assert "query_params" in items_body["query"] or "ItemsQuery" in items_body["query"]
        rule = items_body["variables"]["qp"]["rules"][0]
        assert rule["column_id"] == "status"
        assert rule["compare_value"] == [1]

    def test_group_sugar(self, httpx_mock: HTTPXMock) -> None:
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page([_item("1", "A", {})]),
        )
        result = runner.invoke(
            app,
            ["export", "board", "--board", "42", "-f", "csv", "--group", "topics"],
        )
        assert result.exit_code == 0, result.stdout
        items_body = json.loads(httpx_mock.get_requests()[-1].content)
        rule = items_body["variables"]["qp"]["rules"][0]
        assert rule["column_id"] == "group"
        assert rule["compare_value"] == ["topics"]

    def test_filter_label_resolves_when_column_projected_away(self, httpx_mock: HTTPXMock) -> None:
        # Regression: --columns must not starve --filter of the full column
        # defs. Filtering on `status` while projecting only `Due` must still
        # resolve the "Done" label to its index — sending the raw string makes
        # monday silently return zero matches.
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=_COLS_WITH_STATUS_LABELS)
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page([_item("1", "A", {"status": "Done", "date4": "2026-01-01"})]),
        )
        result = runner.invoke(
            app,
            [
                "export",
                "board",
                "--board",
                "42",
                "-f",
                "csv",
                "--columns",
                "Due",
                "--filter",
                "status=Done",
            ],
        )
        assert result.exit_code == 0, result.stdout
        items_body = json.loads(httpx_mock.get_requests()[-1].content)
        rule = items_body["variables"]["qp"]["rules"][0]
        assert rule["column_id"] == "status"
        assert rule["compare_value"] == [1]  # resolved via full defs, not ["Done"]
        # Projection still applied: the filtered-on Status column is not rendered.
        rows = list(csv.reader(io.StringIO(result.stdout)))
        assert rows[0] == ["id", "name", "state", "group", "Due"]


class TestColumnSubset:
    def test_projection_by_title(self, httpx_mock: HTTPXMock) -> None:
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page([_item("1", "A", {"status": "Done", "date4": "2026-01-01"})]),
        )
        result = runner.invoke(
            app,
            ["export", "board", "--board", "42", "-f", "csv", "--columns", "Status"],
        )
        assert result.exit_code == 0, result.stdout
        rows = list(csv.reader(io.StringIO(result.stdout)))
        assert rows[0] == ["id", "name", "state", "group", "Status"]
        assert "Due" not in rows[0]

    def test_projection_by_id_preserves_order(self, httpx_mock: HTTPXMock) -> None:
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page([_item("1", "A", {"status": "Done", "date4": "2026-01-01"})]),
        )
        result = runner.invoke(
            app,
            ["export", "board", "--board", "42", "-f", "csv", "--columns", "date4,status"],
        )
        assert result.exit_code == 0, result.stdout
        rows = list(csv.reader(io.StringIO(result.stdout)))
        assert rows[0] == ["id", "name", "state", "group", "Due", "Status"]

    def test_unknown_token_errors(self, httpx_mock: HTTPXMock) -> None:
        # The columns fetch resolves first; the bad token aborts before any
        # items page is requested.
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        result = runner.invoke(
            app,
            ["export", "board", "--board", "42", "-f", "csv", "--columns", "nope"],
        )
        assert result.exit_code == 2
        assert "nope" in (result.stdout + str(result.exception or ""))


_COLS_DUP_TITLE = _ok(
    {
        "boards": [
            {
                "id": "42",
                "name": "B",
                "columns": [
                    {"id": "text", "title": "Notes", "type": "text", "archived": False},
                    {"id": "text7", "title": "Notes", "type": "text", "archived": False},
                ],
            }
        ]
    }
)


class TestDuplicateNames:
    def test_duplicate_column_titles_keep_both_values(self, httpx_mock: HTTPXMock) -> None:
        # monday allows two columns with the same title; both must survive
        # (rows are keyed by a disambiguated label, not the raw title).
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=_COLS_DUP_TITLE)
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page([_item("1", "A", {"text": "first", "text7": "second"})]),
        )
        result = runner.invoke(app, ["export", "board", "--board", "42", "-f", "csv"])
        assert result.exit_code == 0, result.stdout
        rows = list(csv.reader(io.StringIO(result.stdout)))
        assert rows[0] == ["id", "name", "state", "group", "Notes (text)", "Notes (text7)"]
        assert rows[1][4] == "first"
        assert rows[1][5] == "second"

    def test_columns_selects_all_same_titled(self, httpx_mock: HTTPXMock) -> None:
        # `--columns Notes` must reach both same-titled columns, not just one.
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=_COLS_DUP_TITLE)
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page([_item("1", "A", {"text": "first", "text7": "second"})]),
        )
        result = runner.invoke(
            app, ["export", "board", "--board", "42", "-f", "csv", "--columns", "Notes"]
        )
        assert result.exit_code == 0, result.stdout
        rows = list(csv.reader(io.StringIO(result.stdout)))
        assert rows[0] == ["id", "name", "state", "group", "Notes (text)", "Notes (text7)"]
        assert rows[1][4] == "first"
        assert rows[1][5] == "second"

    def test_meta_named_column_does_not_shadow_meta(self, httpx_mock: HTTPXMock) -> None:
        # A column literally titled "state" must not overwrite the item's
        # lifecycle state; its label is disambiguated to "state (<id>)".
        cols = _ok(
            {
                "boards": [
                    {
                        "id": "42",
                        "name": "B",
                        "columns": [
                            {"id": "text", "title": "state", "type": "text", "archived": False}
                        ],
                    }
                ]
            }
        )
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=cols)
        httpx_mock.add_response(
            url=ENDPOINT, method="POST", json=_items_page([_item("1", "A", {"text": "custom"})])
        )
        result = runner.invoke(app, ["export", "board", "--board", "42", "-f", "csv"])
        assert result.exit_code == 0, result.stdout
        rows = list(csv.reader(io.StringIO(result.stdout)))
        assert rows[0] == ["id", "name", "state", "group", "state (text)"]
        assert rows[1][2] == "active"  # meta state preserved
        assert rows[1][4] == "custom"  # the column's own value

    def test_duplicate_group_titles_render_separate_sections(self, httpx_mock: HTTPXMock) -> None:
        # Two distinct groups sharing a title must stay separate sections
        # (bucketed by group id, not title).
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=COLS_RESPONSE)
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=_board_name("Roadmap"))
        httpx_mock.add_response(
            url=ENDPOINT,
            method="POST",
            json=_items_page(
                [
                    {**_item("1", "A", {}), "group": {"id": "g1", "title": "Sprint"}},
                    {**_item("2", "B", {}), "group": {"id": "g2", "title": "Sprint"}},
                ]
            ),
        )
        result = runner.invoke(app, ["export", "board", "--board", "42", "-f", "md"])
        assert result.exit_code == 0, result.stdout
        assert result.stdout.count("## Sprint") == 2


class TestDisplayValue:
    """mirror / board_relation cells export display_value, not their null text."""

    def test_board_relation_exports_display_value(self, httpx_mock: HTTPXMock) -> None:
        cols = _ok(
            {
                "boards": [
                    {
                        "id": "42",
                        "name": "B",
                        "columns": [
                            {
                                "id": "link",
                                "title": "Linked",
                                "type": "board_relation",
                                "archived": False,
                            },
                            {
                                "id": "mir",
                                "title": "Mirrored",
                                "type": "mirror",
                                "archived": False,
                            },
                        ],
                    }
                ]
            }
        )
        # monday returns null `text` for both types; the readable value lives in
        # display_value (board_relation also carries linked_item_ids).
        item = {
            "id": "1",
            "name": "A",
            "state": "active",
            "group": {"id": "topics", "title": "Topics"},
            "column_values": [
                {
                    "id": "link",
                    "type": "board_relation",
                    "text": None,
                    "value": None,
                    "display_value": "Item X, Item Y",
                    "linked_item_ids": ["101", "102"],
                },
                {
                    "id": "mir",
                    "type": "mirror",
                    "text": None,
                    "value": None,
                    "display_value": "42",
                },
            ],
        }
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=cols)
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=_items_page([item]))
        result = runner.invoke(app, ["export", "board", "--board", "42", "-f", "csv"])
        assert result.exit_code == 0, result.stdout
        rows = list(csv.reader(io.StringIO(result.stdout)))
        assert rows[0] == ["id", "name", "state", "group", "Linked", "Mirrored"]
        assert rows[1][4] == "Item X, Item Y"
        assert rows[1][5] == "42"

    def test_export_stays_text_first_for_round_trip(self, httpx_mock: HTTPXMock) -> None:
        # Round-trip guard: rating/checkbox cells must export monday's raw
        # `text` ("3" / "v"), NOT codec glyphs ("★★★" / "✓") — the CSV
        # export→import cycle re-parses these strings via parse_value.
        cols = _ok(
            {
                "boards": [
                    {
                        "id": "42",
                        "name": "B",
                        "columns": [
                            {"id": "rate", "title": "Rating", "type": "rating", "archived": False},
                            {"id": "done", "title": "Done", "type": "checkbox", "archived": False},
                        ],
                    }
                ]
            }
        )
        item = {
            "id": "1",
            "name": "A",
            "state": "active",
            "group": {"id": "topics", "title": "Topics"},
            "column_values": [
                {"id": "rate", "type": "rating", "text": "3", "value": '{"rating":3}'},
                {"id": "done", "type": "checkbox", "text": "v", "value": '{"checked":"true"}'},
            ],
        }
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=cols)
        httpx_mock.add_response(url=ENDPOINT, method="POST", json=_items_page([item]))
        result = runner.invoke(app, ["export", "board", "--board", "42", "-f", "csv"])
        assert result.exit_code == 0, result.stdout
        rows = list(csv.reader(io.StringIO(result.stdout)))
        assert rows[1][4] == "3"
        assert rows[1][5] == "v"
